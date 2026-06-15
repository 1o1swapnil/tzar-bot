#!/usr/bin/env python3
"""
build_dashboard.py — Interactive Excel dashboard for the source-code security review skill.

Tabs (see SKILL.md §7):
    1. Cover                — engagement metadata, legend, change log
    2. Executive            — KPI cards, severity donut, OWASP bar, top-5 themes
    3. Findings             — full canonical table, filter-ready
    4. SBOM                 — components inventory + risk
    5. Correlations         — SAST<->DAST<->SCA cross-references
    6. Suppressed (FPs)     — every suppressed finding with reason
    7. Accepted Risk        — open findings the business accepts
    8. Remediation Roadmap  — ranked by ROI
    9. Coverage Gaps        — false-negative honesty
   10. Methodology          — tools, rule packs, severity model

Interactivity:
    - Excel Tables (ListObject) for native filter UI
    - Conditional formatting on severity, EPSS, days-open
    - Data validation on `status`, `owner_role`
    - PivotChart + Chart on Executive tab
    - Defined Names (KPI_* and DATA_*) for downstream linking

Modes:
    --mode simplified | detailed | both

Dependencies:
    pip install openpyxl pyyaml
"""
from __future__ import annotations

import argparse
import json
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    import yaml  # type: ignore
except ImportError:
    yaml = None

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference, PieChart
    from openpyxl.chart.label import DataLabelList
    from openpyxl.formatting.rule import (
        CellIsRule, ColorScaleRule, FormulaRule, IconSetRule,
    )
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.utils import get_column_letter
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.protection import SheetProtection
except ImportError:
    print("[error] pip install openpyxl pyyaml", file=sys.stderr)
    sys.exit(2)

# ----------------------------- styling ---------------------------------------

SEV_FILL = {
    "Critical": PatternFill("solid", fgColor="8B0000"),
    "High":     PatternFill("solid", fgColor="D7263D"),
    "Medium":   PatternFill("solid", fgColor="F49D37"),
    "Low":      PatternFill("solid", fgColor="F1C40F"),
    "Info":     PatternFill("solid", fgColor="3498DB"),
}
SEV_FONT = {
    "Critical": Font(color="FFFFFF", bold=True),
    "High":     Font(color="FFFFFF", bold=True),
    "Medium":   Font(color="000000", bold=True),
    "Low":      Font(color="000000"),
    "Info":     Font(color="FFFFFF"),
}
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(border_style="thin", color="9CA3AF")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

STATUS_VALUES = "new,open,accepted-risk,fixed,fp-suspected,fp-confirmed"
OWNER_VALUES = "Developer,DevOps,Platform,Security,Vendor,Unassigned"

SEV_RANK = {"Critical": 4, "High": 3, "Medium": 2, "Low": 1, "Info": 0}

# Effort hours per tag
EFFORT_HOURS = {"XS": 1, "S": 4, "M": 16, "L": 40, "XL": 120, None: 8}


# ----------------------------- I/O -------------------------------------------

def load_findings(path: Path) -> list[dict[str, Any]]:
    findings: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as f:
        for ln in f:
            ln = ln.strip()
            if ln:
                findings.append(json.loads(ln))
    return findings


def load_scope(path: Path | None) -> dict[str, Any]:
    if not path or not path.exists():
        return {}
    text = path.read_text(encoding="utf-8")
    if yaml:
        try:
            return yaml.safe_load(text) or {}
        except Exception:
            pass
    return {"_raw": text}


def load_sbom(path: Path | None) -> list[dict[str, Any]]:
    if not path or not path.exists():
        return []
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return []
    comps: list[dict[str, Any]] = []
    if data.get("bomFormat") == "CycloneDX":
        for c in data.get("components", []):
            lic = ""
            if c.get("licenses"):
                lic = ", ".join((l.get("license") or {}).get("id") or (l.get("license") or {}).get("name") or "" for l in c["licenses"])
            comps.append({
                "purl": c.get("purl"),
                "name": c.get("name"),
                "version": c.get("version"),
                "type": c.get("type"),
                "scope": c.get("scope"),
                "license": lic,
                "publisher": c.get("publisher") or c.get("author"),
                "description": c.get("description"),
            })
    elif data.get("spdxVersion"):
        for p in data.get("packages", []):
            comps.append({
                "purl": (p.get("externalRefs") or [{}])[0].get("referenceLocator"),
                "name": p.get("name"),
                "version": p.get("versionInfo"),
                "type": "library",
                "scope": p.get("primaryPackagePurpose"),
                "license": p.get("licenseDeclared") or p.get("licenseConcluded"),
                "publisher": p.get("supplier"),
                "description": p.get("description"),
            })
    return comps


# ----------------------------- helpers ---------------------------------------

def autosize(ws, min_w=10, max_w=70) -> None:
    widths: dict[str, int] = {}
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value is None:
                continue
            col = get_column_letter(cell.column)
            ln = len(str(cell.value))
            if ln + 2 > widths.get(col, min_w):
                widths[col] = min(max_w, max(min_w, ln + 2))
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def add_table(ws, name: str, ref: str, style: str = "TableStyleMedium2") -> Table:
    t = Table(displayName=name, ref=ref)
    t.tableStyleInfo = TableStyleInfo(
        name=style, showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(t)
    return t


def write_header(ws, row: int, headers: list[str]) -> None:
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = BORDER


def apply_severity_format(ws, severity_col_letter: str, first_row: int, last_row: int) -> None:
    for sev, fill in SEV_FILL.items():
        rng = f"{severity_col_letter}{first_row}:{severity_col_letter}{last_row}"
        rule = CellIsRule(operator="equal", formula=[f'"{sev}"'],
                          fill=fill, font=SEV_FONT[sev], stopIfTrue=False)
        ws.conditional_formatting.add(rng, rule)


# ----------------------------- tab builders ----------------------------------

def build_cover(wb: Workbook, scope: dict[str, Any], findings: list[dict[str, Any]],
                sbom: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Cover", 0)
    ws.sheet_view.showGridLines = False

    title = ws.cell(row=2, column=2, value="Source Code Security Review — Dashboard")
    title.font = Font(size=20, bold=True, color="111827")
    ws.merge_cells("B2:H2")

    sub = ws.cell(row=3, column=2,
                  value=f"Generated {datetime.now(timezone.utc).isoformat()} (UTC)")
    sub.font = Font(italic=True, color="6B7280")
    ws.merge_cells("B3:H3")

    rows = [
        ("Engagement",       scope.get("engagement_name", "")),
        ("Client",           scope.get("client", "")),
        ("Consultant",       scope.get("consultant", "")),
        ("Engagement type",  scope.get("engagement_type", "")),
        ("Scan window",      f"{scope.get('scan_start','')} → {scope.get('scan_end','')}"),
        ("Application class",scope.get("application_class", "")),
        ("Languages",        ", ".join(scope.get("languages") or [])),
        ("Data sensitivity", scope.get("data_sensitivity", "")),
        ("Deployment model", scope.get("deployment_model", "")),
        ("Compliance",       ", ".join(scope.get("compliance") or [])),
        ("Tools used",       ", ".join(sorted({f.get("source_tool","") for f in findings if f.get("source_tool")}))),
        ("Total findings",   len(findings)),
        ("SBOM components",  len(sbom)),
        ("Source code shared?", "No — report-driven assessment"),
    ]
    for i, (k, v) in enumerate(rows, start=5):
        a = ws.cell(row=i, column=2, value=k); a.font = Font(bold=True); a.border = BORDER
        b = ws.cell(row=i, column=3, value=v); b.border = BORDER
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=8)

    # Legend
    lgrow = 5 + len(rows) + 2
    ws.cell(row=lgrow, column=2, value="Severity legend").font = Font(bold=True, size=12)
    for i, sev in enumerate(["Critical", "High", "Medium", "Low", "Info"], start=lgrow + 1):
        c = ws.cell(row=i, column=2, value=sev)
        c.fill = SEV_FILL[sev]; c.font = SEV_FONT[sev]
        c.alignment = Alignment(horizontal="center")
    autosize(ws)
    ws.protection = SheetProtection(sheet=True, selectLockedCells=True, selectUnlockedCells=True)


def build_findings(wb: Workbook, findings: list[dict[str, Any]], mode: str,
                   sheet_title: str | None = None, table_name: str | None = None) -> str:
    ws = wb.create_sheet(sheet_title or "Findings")
    tname = table_name or "tbl_findings"
    detailed_cols = [
        "finding_id", "severity_normalized", "title", "cwe", "owasp_top10_2021",
        "source_tool", "evidence_tools", "source_rule_id",
        "language", "asset", "component_purl", "component_version", "fixed_in_version",
        "location_file", "location_start_line", "location_end_line", "location_url",
        "cvss_v3_1", "epss", "kev_listed",
        "status", "owner_role", "remediation_effort",
        "fp_reasoning", "remediation", "description", "evidence",
        "first_seen", "last_seen", "tags",
    ]
    simplified_cols = [
        "severity_normalized", "title", "asset", "component_purl",
        "status", "owner_role", "remediation_effort", "remediation",
    ]
    cols = simplified_cols if mode == "simplified" else detailed_cols

    write_header(ws, 1, cols)
    for r, f in enumerate(findings, start=2):
        for c, key in enumerate(cols, start=1):
            v = f.get(key)
            if isinstance(v, list):
                v = ", ".join(map(str, v))
            ws.cell(row=r, column=c, value=v)

    last_row = max(2, 1 + len(findings))
    last_col_letter = get_column_letter(len(cols))
    ref = f"A1:{last_col_letter}{last_row}"
    add_table(ws, tname, ref)

    # severity formatting
    sev_idx = cols.index("severity_normalized") + 1
    apply_severity_format(ws, get_column_letter(sev_idx), 2, last_row)

    # status & owner validation
    if "status" in cols:
        dv = DataValidation(type="list", formula1=f'"{STATUS_VALUES}"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{get_column_letter(cols.index('status')+1)}2:{get_column_letter(cols.index('status')+1)}{last_row}")
    if "owner_role" in cols:
        dv2 = DataValidation(type="list", formula1=f'"{OWNER_VALUES}"', allow_blank=True)
        ws.add_data_validation(dv2)
        dv2.add(f"{get_column_letter(cols.index('owner_role')+1)}2:{get_column_letter(cols.index('owner_role')+1)}{last_row}")

    # EPSS gradient if present
    if "epss" in cols:
        epss_col = get_column_letter(cols.index("epss") + 1)
        ws.conditional_formatting.add(
            f"{epss_col}2:{epss_col}{last_row}",
            ColorScaleRule(start_type="num", start_value=0, start_color="63BE7B",
                           mid_type="num", mid_value=0.5, mid_color="FFEB84",
                           end_type="num", end_value=1, end_color="F8696B"))

    # KEV flag
    if "kev_listed" in cols:
        kev_col = get_column_letter(cols.index("kev_listed") + 1)
        ws.conditional_formatting.add(
            f"{kev_col}2:{kev_col}{last_row}",
            CellIsRule(operator="equal", formula=["TRUE"],
                       fill=PatternFill("solid", fgColor="8B0000"),
                       font=Font(color="FFFFFF", bold=True), stopIfTrue=False))

    autosize(ws, max_w=60)
    ws.freeze_panes = "C2"
    return ref


def build_executive(wb: Workbook, findings: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Executive", 1)
    ws.sheet_view.showGridLines = False

    sev_counter = Counter((f.get("severity_normalized") or "Info") for f in findings)
    total = sum(sev_counter.values())
    critical = sev_counter.get("Critical", 0)
    high = sev_counter.get("High", 0)
    med = sev_counter.get("Medium", 0)
    low = sev_counter.get("Low", 0)

    owasp_counter: Counter = Counter()
    for f in findings:
        for o in f.get("owasp_top10_2021") or []:
            owasp_counter[o] += 1

    open_findings = [f for f in findings if (f.get("status") or "new") not in ("fixed", "fp-confirmed")]
    fp_count = sum(1 for f in findings if (f.get("status") or "") == "fp-confirmed")
    accepted = sum(1 for f in findings if (f.get("status") or "") == "accepted-risk")

    # KPI cards
    cards = [
        ("Total findings", total),
        ("Critical", critical),
        ("High", high),
        ("Medium", med),
        ("Open", len(open_findings)),
        ("Suppressed (FP)", fp_count),
        ("Accepted risk", accepted),
    ]
    ws.cell(row=1, column=1, value="Executive Summary").font = Font(size=18, bold=True)
    for i, (label, val) in enumerate(cards, start=0):
        col = 1 + (i * 2)
        c1 = ws.cell(row=3, column=col, value=label)
        c1.font = Font(bold=True, color="6B7280"); c1.alignment = Alignment(horizontal="center")
        c2 = ws.cell(row=4, column=col, value=val)
        c2.font = Font(size=22, bold=True, color="111827")
        c2.alignment = Alignment(horizontal="center")
        ws.cell(row=4, column=col).border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = 16

    # Severity table for chart
    sev_order = ["Critical", "High", "Medium", "Low", "Info"]
    ws.cell(row=8, column=1, value="Severity").font = Font(bold=True)
    ws.cell(row=8, column=2, value="Count").font = Font(bold=True)
    for i, s in enumerate(sev_order, start=9):
        ws.cell(row=i, column=1, value=s)
        ws.cell(row=i, column=2, value=sev_counter.get(s, 0))

    donut = DoughnutChart()
    donut.title = "Findings by Severity"
    donut.add_data(Reference(ws, min_col=2, min_row=8, max_row=8 + len(sev_order)), titles_from_data=True)
    donut.set_categories(Reference(ws, min_col=1, min_row=9, max_row=8 + len(sev_order)))
    donut.dataLabels = DataLabelList(showPercent=True)
    ws.add_chart(donut, "D8")

    # OWASP bar
    ws.cell(row=8, column=12, value="OWASP 2021").font = Font(bold=True)
    ws.cell(row=8, column=13, value="Count").font = Font(bold=True)
    for i, (k, v) in enumerate(sorted(owasp_counter.items()), start=9):
        ws.cell(row=i, column=12, value=k)
        ws.cell(row=i, column=13, value=v)
    if owasp_counter:
        bar = BarChart()
        bar.type = "bar"
        bar.title = "Findings by OWASP 2021"
        bar.add_data(Reference(ws, min_col=13, min_row=8, max_row=8 + len(owasp_counter)), titles_from_data=True)
        bar.set_categories(Reference(ws, min_col=12, min_row=9, max_row=8 + len(owasp_counter)))
        ws.add_chart(bar, "O8")

    # Defined names for downstream linking
    wb.defined_names["KPI_total"] = DefinedName(name="KPI_total", attr_text=f"Executive!$B$4")
    wb.defined_names["KPI_critical"] = DefinedName(name="KPI_critical", attr_text=f"Executive!$D$4")
    wb.defined_names["KPI_high"] = DefinedName(name="KPI_high", attr_text=f"Executive!$F$4")
    wb.defined_names["KPI_open"] = DefinedName(name="KPI_open", attr_text=f"Executive!$J$4")


def build_sbom(wb: Workbook, sbom: list[dict[str, Any]], findings: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("SBOM")
    cols = ["purl", "name", "version", "type", "scope", "license", "publisher",
            "open_vulns", "max_severity", "kev_in_pkg", "description"]
    write_header(ws, 1, cols)

    # Build a vuln index from findings
    vuln_by_purl: dict[str, list[dict[str, Any]]] = {}
    for f in findings:
        p = f.get("component_purl")
        if p:
            vuln_by_purl.setdefault(p, []).append(f)

    for r, comp in enumerate(sbom, start=2):
        vulns = vuln_by_purl.get(comp.get("purl"), [])
        max_sev = max((v.get("severity_normalized") for v in vulns), key=lambda s: SEV_RANK.get(s, -1), default=None)
        kev = any(v.get("kev_listed") for v in vulns)
        values = [
            comp.get("purl"), comp.get("name"), comp.get("version"),
            comp.get("type"), comp.get("scope"), comp.get("license"),
            comp.get("publisher"), len(vulns), max_sev, kev, comp.get("description"),
        ]
        for c, v in enumerate(values, start=1):
            ws.cell(row=r, column=c, value=v)

    last_row = max(2, 1 + len(sbom))
    if sbom:
        ref = f"A1:{get_column_letter(len(cols))}{last_row}"
        add_table(ws, "tbl_sbom", ref, style="TableStyleMedium4")

        sev_col = get_column_letter(cols.index("max_severity") + 1)
        apply_severity_format(ws, sev_col, 2, last_row)

        kev_col = get_column_letter(cols.index("kev_in_pkg") + 1)
        ws.conditional_formatting.add(
            f"{kev_col}2:{kev_col}{last_row}",
            CellIsRule(operator="equal", formula=["TRUE"],
                       fill=PatternFill("solid", fgColor="8B0000"),
                       font=Font(color="FFFFFF", bold=True), stopIfTrue=False))

    autosize(ws, max_w=60)
    ws.freeze_panes = "C2"


def build_correlations(wb: Workbook, findings: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Correlations")
    cols = ["finding_id", "title", "source_tool", "asset", "location_url",
            "severity_normalized", "correlated_with"]
    write_header(ws, 1, cols)
    r = 2
    for f in findings:
        corr = f.get("correlations") or []
        if not corr:
            continue
        ws.cell(row=r, column=1, value=f.get("finding_id"))
        ws.cell(row=r, column=2, value=f.get("title"))
        ws.cell(row=r, column=3, value=f.get("source_tool"))
        ws.cell(row=r, column=4, value=f.get("asset"))
        ws.cell(row=r, column=5, value=f.get("location_url"))
        ws.cell(row=r, column=6, value=f.get("severity_normalized"))
        ws.cell(row=r, column=7, value=", ".join(corr))
        r += 1
    if r > 2:
        add_table(ws, "tbl_correlations", f"A1:{get_column_letter(len(cols))}{r-1}", style="TableStyleMedium6")
        apply_severity_format(ws, get_column_letter(cols.index("severity_normalized")+1), 2, r-1)
    autosize(ws, max_w=70)


def build_suppressed(wb: Workbook, findings: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Suppressed (FPs)")
    cols = ["finding_id", "severity_normalized", "title", "source_tool",
            "source_rule_id", "asset", "location_file", "location_start_line",
            "fp_reasoning", "approver", "approved_at"]
    write_header(ws, 1, cols)
    suppressed = [f for f in findings if (f.get("status") or "") == "fp-confirmed"]
    for r, f in enumerate(suppressed, start=2):
        for c, key in enumerate(cols, start=1):
            ws.cell(row=r, column=c, value=f.get(key) if key in f else None)
    if suppressed:
        add_table(ws, "tbl_suppressed", f"A1:{get_column_letter(len(cols))}{1+len(suppressed)}",
                  style="TableStyleMedium11")
        apply_severity_format(ws, get_column_letter(cols.index("severity_normalized")+1), 2, 1+len(suppressed))
    autosize(ws, max_w=70)
    ws.protection = SheetProtection(sheet=True, selectLockedCells=True, selectUnlockedCells=True)


def build_accepted(wb: Workbook, findings: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Accepted Risk")
    cols = ["finding_id", "severity_normalized", "title", "asset", "owner_role",
            "compensating_controls", "expiry_date", "approver"]
    write_header(ws, 1, cols)
    accepted = [f for f in findings if (f.get("status") or "") == "accepted-risk"]
    for r, f in enumerate(accepted, start=2):
        for c, key in enumerate(cols, start=1):
            ws.cell(row=r, column=c, value=f.get(key) if key in f else None)
    if accepted:
        add_table(ws, "tbl_accepted", f"A1:{get_column_letter(len(cols))}{1+len(accepted)}",
                  style="TableStyleMedium9")
        apply_severity_format(ws, get_column_letter(cols.index("severity_normalized")+1), 2, 1+len(accepted))
    autosize(ws, max_w=60)


def compute_roi(f: dict[str, Any]) -> float:
    sev = {"Critical": 10, "High": 7, "Medium": 4, "Low": 1, "Info": 0}.get(f.get("severity_normalized"), 0)
    mult = 1.0
    if f.get("kev_listed"): mult *= 1.5
    epss = f.get("epss")
    if epss and epss >= 0.7: mult *= 1.3
    hours = EFFORT_HOURS.get(f.get("remediation_effort"), 8)
    return round(sev * mult / max(1, hours), 3)


def build_roadmap(wb: Workbook, findings: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Remediation Roadmap")
    cols = ["roi", "severity_normalized", "title", "asset", "component_purl",
            "remediation", "remediation_effort", "owner_role", "status",
            "sprint", "target_date", "finding_id"]
    write_header(ws, 1, cols)
    open_findings = [f for f in findings if (f.get("status") or "") not in ("fixed", "fp-confirmed")]
    open_findings.sort(key=compute_roi, reverse=True)
    for r, f in enumerate(open_findings, start=2):
        ws.cell(row=r, column=1, value=compute_roi(f))
        ws.cell(row=r, column=2, value=f.get("severity_normalized"))
        ws.cell(row=r, column=3, value=f.get("title"))
        ws.cell(row=r, column=4, value=f.get("asset"))
        ws.cell(row=r, column=5, value=f.get("component_purl"))
        ws.cell(row=r, column=6, value=f.get("remediation"))
        ws.cell(row=r, column=7, value=f.get("remediation_effort"))
        ws.cell(row=r, column=8, value=f.get("owner_role"))
        ws.cell(row=r, column=9, value=f.get("status"))
        ws.cell(row=r, column=10, value=None)
        ws.cell(row=r, column=11, value=None)
        ws.cell(row=r, column=12, value=f.get("finding_id"))
    if open_findings:
        add_table(ws, "tbl_roadmap",
                  f"A1:{get_column_letter(len(cols))}{1+len(open_findings)}",
                  style="TableStyleMedium7")
        apply_severity_format(ws, get_column_letter(cols.index("severity_normalized")+1),
                              2, 1+len(open_findings))
        # ROI gradient
        roi_col = "A"
        ws.conditional_formatting.add(
            f"{roi_col}2:{roi_col}{1+len(open_findings)}",
            ColorScaleRule(start_type="min", start_color="FFFFFF",
                           end_type="max", end_color="2ECC71"))
    autosize(ws, max_w=60)
    ws.freeze_panes = "C2"


def build_gaps(wb: Workbook, findings: list[dict[str, Any]], scope: dict[str, Any]) -> None:
    ws = wb.create_sheet("Coverage Gaps")
    rows = [
        ("Tool", "Languages covered", "Languages NOT covered", "Notes"),
    ]
    expected_langs = set(scope.get("languages") or [])
    seen_per_tool: dict[str, set[str]] = {}
    for f in findings:
        tool = f.get("source_tool") or "?"
        lang = (f.get("language") or "").strip()
        if lang:
            seen_per_tool.setdefault(tool, set()).add(lang)
    write_header(ws, 1, list(rows[0]))
    r = 2
    for tool, langs in sorted(seen_per_tool.items()):
        not_covered = ", ".join(sorted(expected_langs - langs)) if expected_langs else "(scope languages not provided)"
        ws.cell(row=r, column=1, value=tool)
        ws.cell(row=r, column=2, value=", ".join(sorted(langs)))
        ws.cell(row=r, column=3, value=not_covered)
        ws.cell(row=r, column=4, value="")
        r += 1
    if r > 2:
        add_table(ws, "tbl_gaps", f"A1:D{r-1}", style="TableStyleMedium8")
    autosize(ws, max_w=60)


def build_methodology(wb: Workbook, findings: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Methodology")
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value="Methodology").font = Font(size=18, bold=True)
    paragraphs = [
        ("Scan inputs",
         "This dashboard was generated exclusively from client-provided scanner outputs. "
         "No raw source code was reviewed."),
        ("Normalization",
         "All scanner outputs are mapped to a canonical Finding schema with deterministic IDs."),
        ("Severity model",
         "CVSS bands first; vendor severities as fallback; KEV floors at Critical; EPSS≥0.7 adds one level."),
        ("Deduplication",
         "Strong match: CWE+file+line±2+language. SCA match: PURL+CVE. DAST↔SAST kept as correlations, not merged."),
        ("False positives",
         "Suppression requires written justification — captured on the Suppressed tab. Defense-in-depth claims do not qualify."),
        ("Coverage caveats",
         "Findings outside the scanners' rule packs / paths are out of scope. See Coverage Gaps."),
        ("Sign-off",
         "Acceptance of suppressions and accepted-risk decisions is the client's. The consultant attests only to the analysis of the inputs provided."),
    ]
    r = 3
    for h, body in paragraphs:
        a = ws.cell(row=r, column=1, value=h); a.font = Font(bold=True, size=12)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        b = ws.cell(row=r+1, column=1, value=body)
        b.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=8)
        ws.row_dimensions[r+1].height = 50
        r += 3
    # Tool inventory
    ws.cell(row=r, column=1, value="Tool inventory").font = Font(bold=True, size=12)
    r += 1
    write_header(ws, r, ["Tool", "Versions seen", "Findings"])
    tool_counter: Counter = Counter()
    versions: dict[str, set[str]] = {}
    for f in findings:
        t = f.get("source_tool") or "?"
        tool_counter[t] += 1
        if f.get("source_tool_version"):
            versions.setdefault(t, set()).add(str(f["source_tool_version"]))
    r += 1
    for tool, cnt in tool_counter.most_common():
        ws.cell(row=r, column=1, value=tool)
        ws.cell(row=r, column=2, value=", ".join(sorted(versions.get(tool, set()))) or "—")
        ws.cell(row=r, column=3, value=cnt)
        r += 1
    autosize(ws, max_w=80)
    ws.protection = SheetProtection(sheet=True, selectLockedCells=True, selectUnlockedCells=True)


# ----------------------------- entry -----------------------------------------

def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--findings", required=True, type=Path)
    ap.add_argument("--sbom", type=Path)
    ap.add_argument("--scope", type=Path)
    ap.add_argument("--mode", choices=["simplified", "detailed", "both"], default="both")
    ap.add_argument("--out", default="dashboard.xlsx")
    args = ap.parse_args()

    findings = load_findings(args.findings)
    scope = load_scope(args.scope)
    sbom = load_sbom(args.sbom)

    wb = Workbook()
    # remove default
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    build_cover(wb, scope, findings, sbom)
    build_executive(wb, findings)

    if args.mode == "simplified":
        build_findings(wb, findings, "simplified",
                       sheet_title="Findings", table_name="tbl_findings")
    elif args.mode == "detailed":
        build_findings(wb, findings, "detailed",
                       sheet_title="Findings", table_name="tbl_findings")
    else:  # both
        build_findings(wb, findings, "simplified",
                       sheet_title="Findings (Simplified)",
                       table_name="tbl_findings_simple")
        build_findings(wb, findings, "detailed",
                       sheet_title="Findings (Detailed)",
                       table_name="tbl_findings_detail")

    build_sbom(wb, sbom, findings)
    build_correlations(wb, findings)
    build_suppressed(wb, findings)
    build_accepted(wb, findings)
    build_roadmap(wb, findings)
    build_gaps(wb, findings, scope)
    build_methodology(wb, findings)

    wb.save(args.out)
    print(f"[ok] wrote {args.out} ({len(findings)} findings, {len(sbom)} components)", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
